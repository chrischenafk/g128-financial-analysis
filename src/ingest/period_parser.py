"""Period parser — the first real validation gate of the ingest layer.

Single job: turn ONE scanned ``.xlsm`` path into a validated, structured
description of its two reporting periods. It does three things, in order, and
fails loud at the first sign of trouble:

  1. Parse the filename's ``<YYYY>_<MM>_vs_<YYYY>_<MM>`` pattern into two
     periods. The chronologically LATER period is ``current``; the earlier is
     ``comparison``. Filename order is not trusted — order is computed.
  2. Classify the pair as ``MoM`` (1-month gap) or ``YoY`` (12-month gap) from
     the gap itself, never from a filename label. Any other gap is an error.
  3. Cross-check against the workbook: open it read-only, find the Summary
     sheet, scan its cells for ``MM/DD/YYYY - MM/DD/YYYY`` period headers, and
     assert BOTH filename periods appear there. Filenames must agree with the
     in-workbook headers (AGENTS.md §4) — a mismatch is information, not a thing
     to smooth over.

Scope boundary: this module reads HEADER CELLS ONLY for the cross-check. It does
not load full sheets, build DataFrames, or compute any metric — that belongs to
``excel_loader.py``. Cross-file MoM/YoY pairing and the anchor-match between two
files is also out of scope here (a later ingest step).
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.utils.logger import get_logger

logger = get_logger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────
# The Summary sheet name is TikTok-specific. Fine for current scope (TikTok Shop
# only); when other marketplaces arrive this becomes a per-marketplace setting.
SUMMARY_SHEET_NAME = "TikTok Summary"

# Period pattern in the filename stem: <YYYY><sep><MM> vs <YYYY><sep><MM>.
# We are strict about the four-number structure but tolerant of the separators,
# because real workbooks arrive named "...2025.04 vs 2026.04.xlsm" (dot between
# year/month, spaces around "vs") while other sources use underscores
# ("..._2025_04_vs_2026_04.xlsm"). Year/month joiner may be '.', '_' or '-'; the
# "vs" may be flanked by underscores or whitespace. We search anywhere in the
# stem so the varying prefix (Tiktok SKU-Level Profit ..., etc.) is ignored.
_FILENAME_PERIOD_RE = re.compile(
    r"(\d{4})[._-](\d{2})[_\s]*vs[_\s]*(\d{4})[._-](\d{2})"
)

# In-workbook period header, e.g. "03/01/2026 - 03/31/2026". We only need the
# START date of each range to derive the period (year, month). The layout is
# discovered by scanning cells, never hardcoded to a coordinate.
_HEADER_PERIOD_RE = re.compile(
    r"(\d{2})/(\d{2})/(\d{4})\s*-\s*(\d{2})/(\d{2})/(\d{4})"
)

# Gap (in months) → comparison type. Anything not in this map is unsupported.
_GAP_TO_TYPE = {1: "MoM", 12: "YoY"}


# ─────────────────────────────────────────────────────────────────────────────
# Data structures
# ─────────────────────────────────────────────────────────────────────────────
@dataclass(frozen=True, order=True)
class Period:
    """A single reporting month.

    ``order=True`` gives natural chronological comparison/sorting: the fields are
    compared in declaration order, so (year, month) sorts correctly without a
    custom key. ``frozen=True`` makes it hashable and safe to use in sets/dicts
    (we compare periods for membership during the cross-check).
    """

    year: int
    month: int

    def __post_init__(self) -> None:
        # Guard the invariant at construction so an impossible month can never
        # silently flow downstream. (1..12 only — calendar months.)
        if not 1 <= self.month <= 12:
            raise ValueError(
                f"Impossible month {self.month:02d} for period "
                f"{self.year}-{self.month:02d}: month must be 01–12."
            )

    def __str__(self) -> str:  # readable "2026-04"
        return f"{self.year:04d}-{self.month:02d}"


@dataclass(frozen=True)
class FilePeriods:
    """The validated period description for one workbook file.

    ``current`` is the chronologically later period, ``comparison`` the earlier.
    ``comparison_type`` is "MoM" or "YoY", inferred from the gap.
    """

    path: Path
    current: Period
    comparison: Period
    comparison_type: str

# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — filename parsing
# ─────────────────────────────────────────────────────────────────────────────
def parse_filename_periods(path: Path) -> tuple[Period, Period]:
    """Extract (current, comparison) periods from the filename stem.

    ``current`` = the chronologically later of the two periods, ``comparison`` =
    the earlier. Filename ordering is NOT assumed; we sort the two periods.

    Raises:
        ValueError: if the strict period pattern is absent, or a month is
            impossible (00, 13+). The offending filename is named in the message.
    """
    match = _FILENAME_PERIOD_RE.search(path.stem)
    if match is None:
        raise ValueError(
            f"Filename does not contain the required period pattern "
            f"'<YYYY>_<MM>_vs_<YYYY>_<MM>': {path.name!r}. Never guessing a "
            "period from a malformed name."
        )

    y1, m1, y2, m2 = (int(g) for g in match.groups())
    try:
        first = Period(y1, m1)
        second = Period(y2, m2)
    except ValueError as exc:  # impossible month — re-raise naming the file
        raise ValueError(
            f"Filename {path.name!r} contains an impossible month: {exc}"
        ) from exc

    current, comparison = max(first, second), min(first, second)
    return current, comparison


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — MoM vs YoY classification (from the gap, not a label)
# ─────────────────────────────────────────────────────────────────────────────
def classify_comparison(current: Period, comparison: Period) -> str:
    """Return "MoM" (gap 1) or "YoY" (gap 12) from the months-between gap.

    Raises:
        ValueError: for any other gap (including 0 — the same period twice).
            We never silently accept a 2- or 6-month gap as either lens.
    """
    months_between = (current.year - comparison.year) * 12 + (
        current.month - comparison.month
    )
    comparison_type = _GAP_TO_TYPE.get(months_between)
    if comparison_type is None:
        raise ValueError(
            f"Unsupported period gap of {months_between} month(s) between "
            f"current {current} and comparison {comparison}. Only a 1-month gap "
            "(MoM) or 12-month gap (YoY) is supported; refusing to guess."
        )
    return comparison_type


# ─────────────────────────────────────────────────────────────────────────────
# Step 3 — workbook cross-check (the validation gate)
# ─────────────────────────────────────────────────────────────────────────────
def read_summary_periods(path: Path) -> set[Period]:
    """Open the workbook read-only and return the periods in its Summary headers.

    Reads HEADER CELLS ONLY — it scans the Summary sheet's cell values for the
    ``MM/DD/YYYY - MM/DD/YYYY`` pattern and parses each match's START date into a
    Period. It does not load data or build DataFrames.

    The workbook is always closed (try/finally) — Windows holds a file lock on an
    open ``.xlsm`` and a leaked handle would block later steps.

    Raises:
        ValueError: if the Summary sheet is absent.
    """
    # data_only=True so we read computed values, not formula strings; read_only
    # streams cells without loading the whole workbook into memory.
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if SUMMARY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Workbook {path.name!r} is missing the required "
                f"{SUMMARY_SHEET_NAME!r} sheet. Found sheets: "
                f"{workbook.sheetnames}."
            )

        sheet = workbook[SUMMARY_SHEET_NAME]
        periods: set[Period] = set()
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if not isinstance(value, str):
                    continue
                for header in _HEADER_PERIOD_RE.finditer(value):
                    # groups: start MM, DD, YYYY, end MM, DD, YYYY — START defines
                    # the period (a month range starts in its own month).
                    start_month = int(header.group(1))
                    start_year = int(header.group(3))
                    periods.add(Period(start_year, start_month))
        return periods
    finally:
        workbook.close()


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────
def parse_and_validate(path: Path) -> FilePeriods:
    """Parse the filename, classify the gap, and cross-check the workbook.

    Runs filename-parse → classify → workbook cross-check and returns the
    validated ``FilePeriods``. Any disagreement (bad filename, unsupported gap,
    missing sheet, or filename periods absent from the workbook) raises a clear
    error and stops — never a silent guess.
    """
    current, comparison = parse_filename_periods(path)
    comparison_type = classify_comparison(current, comparison)

    workbook_periods = read_summary_periods(path)
    missing = [p for p in (current, comparison) if p not in workbook_periods]
    if missing:
        found = sorted(workbook_periods)
        raise ValueError(
            f"Filename/workbook period mismatch in {path.name!r}: the filename "
            f"claims current={current}, comparison={comparison}, but the "
            f"{SUMMARY_SHEET_NAME!r} headers only contain "
            f"{[str(p) for p in found] or '[]'}. Missing: "
            f"{[str(p) for p in missing]}. Filenames must agree with in-workbook "
            "headers."
        )

    result = FilePeriods(
        path=path,
        current=current,
        comparison=comparison,
        comparison_type=comparison_type,
    )
    logger.info(
        "Validated %s: current=%s, comparison=%s, type=%s.",
        path.name,
        current,
        comparison,
        comparison_type,
    )
    return result
