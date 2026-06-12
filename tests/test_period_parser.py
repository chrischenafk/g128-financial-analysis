"""Tests for src/ingest/period_parser.py.

Two layers, both using only synthetic data (tmp_path + openpyxl-built
workbooks) — never real company data:

  1. Pure filename parsing & gap classification (reads the name only): MoM/YoY
     period extraction, ordering independence, and the loud failures (no period
     pattern, impossible month, unsupported/zero gaps).
  2. The workbook cross-check: build a minimal "TikTok Summary" sheet carrying
     the expected date-range headers and assert parse_and_validate succeeds;
     then negative cases — headers that disagree, and a missing Summary sheet.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from src.ingest.period_parser import (
    SUMMARY_SHEET_NAME,
    FilePeriods,
    Period,
    classify_comparison,
    parse_and_validate,
    parse_filename_periods,
)

# Real-convention filenames (the exact lowercase-t, underscore style from §5).
MOM_NAME = "Tiktok_SKULevel_Profit_2026_03_vs_2026_04.xlsm"
YOY_NAME = "Tiktok_SKULevel_Profit_2025_04_vs_2026_04.xlsm"


# ── helpers ──────────────────────────────────────────────────────────────────
def _make_workbook(
    path: Path, headers: list[str], sheet_name: str = SUMMARY_SHEET_NAME
) -> Path:
    """Write a minimal workbook with one sheet containing the given header
    strings (one per column in row 1, mimicking the periods-as-columns layout).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Period headers live somewhere in the sheet; put them across row 1 starting
    # at column B (column A is a line-item label), to prove discovery-not-coords.
    ws.cell(row=1, column=1, value="Line Item")
    for offset, header in enumerate(headers):
        ws.cell(row=1, column=2 + offset, value=header)
    ws.cell(row=2, column=1, value="Gross")  # a non-period cell, harmlessly ignored
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — pure filename parsing
# ─────────────────────────────────────────────────────────────────────────────
def test_period_str_and_ordering() -> None:
    assert str(Period(2026, 4)) == "2026-04"
    assert Period(2026, 3) < Period(2026, 4)
    assert Period(2025, 12) < Period(2026, 1)
    assert max(Period(2026, 4), Period(2026, 3)) == Period(2026, 4)


def test_filename_mom_periods_and_order() -> None:
    current, comparison = parse_filename_periods(Path(MOM_NAME))
    assert current == Period(2026, 4)
    assert comparison == Period(2026, 3)
    assert classify_comparison(current, comparison) == "MoM"


def test_filename_yoy_periods_and_order() -> None:
    current, comparison = parse_filename_periods(Path(YOY_NAME))
    assert current == Period(2026, 4)
    assert comparison == Period(2025, 4)
    assert classify_comparison(current, comparison) == "YoY"


def test_filename_order_is_computed_not_assumed() -> None:
    # Later period listed FIRST in the name — current must still be the later one.
    current, comparison = parse_filename_periods(
        Path("Tiktok_SKULevel_Profit_2026_04_vs_2026_03.xlsm")
    )
    assert current == Period(2026, 4)
    assert comparison == Period(2026, 3)


def test_filename_without_period_pattern_raises() -> None:
    with pytest.raises(ValueError, match="period pattern"):
        parse_filename_periods(Path("Tiktok_SKULevel_Profit_April.xlsm"))


def test_filename_impossible_month_raises() -> None:
    with pytest.raises(ValueError, match="month"):
        parse_filename_periods(Path("Tiktok_2026_13_vs_2026_04.xlsm"))


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — gap classification failures
# ─────────────────────────────────────────────────────────────────────────────
def test_zero_gap_raises() -> None:
    # Same period twice → gap 0 → neither MoM nor YoY.
    with pytest.raises(ValueError, match="0 month"):
        classify_comparison(Period(2026, 4), Period(2026, 4))


def test_two_month_gap_raises() -> None:
    with pytest.raises(ValueError, match="2 month"):
        classify_comparison(Period(2026, 4), Period(2026, 2))


def test_six_month_gap_raises() -> None:
    with pytest.raises(ValueError, match="6 month"):
        classify_comparison(Period(2026, 4), Period(2025, 10))


# ─────────────────────────────────────────────────────────────────────────────
# Step 3 — workbook cross-check
# ─────────────────────────────────────────────────────────────────────────────
def test_parse_and_validate_mom_success(tmp_path: Path) -> None:
    path = tmp_path / MOM_NAME
    _make_workbook(
        path,
        headers=["03/01/2026 - 03/31/2026", "04/01/2026 - 04/30/2026"],
    )
    result = parse_and_validate(path)
    assert result == FilePeriods(
        path=path,
        current=Period(2026, 4),
        comparison=Period(2026, 3),
        comparison_type="MoM",
    )


def test_parse_and_validate_yoy_success(tmp_path: Path) -> None:
    path = tmp_path / YOY_NAME
    _make_workbook(
        path,
        headers=["04/01/2025 - 04/30/2025", "04/01/2026 - 04/30/2026"],
    )
    result = parse_and_validate(path)
    assert result.current == Period(2026, 4)
    assert result.comparison == Period(2025, 4)
    assert result.comparison_type == "YoY"


def test_cross_check_mismatch_raises(tmp_path: Path) -> None:
    # Filename claims 2026-03 vs 2026-04, but the workbook headers are a
    # different (YoY-shaped) pair → the gate must fire.
    path = tmp_path / MOM_NAME
    _make_workbook(
        path,
        headers=["04/01/2025 - 04/30/2025", "04/01/2026 - 04/30/2026"],
    )
    with pytest.raises(ValueError, match="mismatch"):
        parse_and_validate(path)


def test_cross_check_partial_mismatch_raises(tmp_path: Path) -> None:
    # Current period present, comparison period absent → still a mismatch.
    path = tmp_path / MOM_NAME
    _make_workbook(
        path,
        headers=["04/01/2026 - 04/30/2026", "01/01/2026 - 01/31/2026"],
    )
    with pytest.raises(ValueError, match="Missing"):
        parse_and_validate(path)


def test_missing_summary_sheet_raises(tmp_path: Path) -> None:
    path = tmp_path / MOM_NAME
    _make_workbook(
        path,
        headers=["03/01/2026 - 03/31/2026", "04/01/2026 - 04/30/2026"],
        sheet_name="Some Other Sheet",
    )
    with pytest.raises(ValueError, match="missing the required"):
        parse_and_validate(path)
