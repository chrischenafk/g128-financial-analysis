"""Tests for src/ingest/excel_loader.py.

Synthetic tmp_path workbooks built with openpyxl — no real company data. The
loader's contract: load present sheets faithfully (no rows dropped or altered),
discover the Profit Margin header row even when it isn't the first row, require
the two core sheets, and degrade gracefully (None + warning) for absent optional
data-quality sheets.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from src.ingest.excel_loader import (
    PROFIT_MARGIN_SHEET,
    SUMMARY_SHEET,
    UNMAPPED_ADS_SHEET,
    load_workbook,
)

# Profit Margin rows the loader must preserve verbatim, including a zero/blank
# row that must survive intact (no dedupe, no NaN-fill, no active-SKU filter).
_PM_HEADER = ["Marketplace SKU", "Date Range", "Units", "Total Gross Sale"]
_PM_ROWS = [
    ["SKU-A", "04/01/2026 - 04/30/2026", 10, 100.0],
    ["SKU-A", "03/01/2026 - 03/31/2026", 8, 80.0],
    ["SKU-B", "04/01/2026 - 04/30/2026", 0, 0.0],  # zero-activity row — must stay
    ["SKU-B", "03/01/2026 - 03/31/2026", None, None],  # blank values — must stay
]


def _build_workbook(
    path: Path,
    *,
    pm_header_offset: int = 0,
    include_profit_margin: bool = True,
    include_unmapped_ads: bool = True,
) -> Path:
    """Write a synthetic workbook to ``path``.

    ``pm_header_offset`` inserts that many junk/title rows above the Profit
    Margin header, directly exercising header discovery.
    """
    wb = Workbook()

    # Summary sheet (periods as columns, line items as rows).
    summary = wb.active
    summary.title = SUMMARY_SHEET
    summary.append(["Line Item", "03/01/2026 - 03/31/2026", "04/01/2026 - 04/30/2026"])
    summary.append(["Gross", 80.0, 100.0])
    summary.append(["Profit", 20.0, 25.0])

    if include_profit_margin:
        pm = wb.create_sheet(PROFIT_MARGIN_SHEET)
        for i in range(pm_header_offset):
            pm.append([f"-- report title row {i} --"])  # junk above the header
        pm.append(_PM_HEADER)
        for row in _PM_ROWS:
            pm.append(row)

    if include_unmapped_ads:
        ads = wb.create_sheet(UNMAPPED_ADS_SHEET)
        ads.append(["Ad ID", "Spend"])
        ads.append(["AD-1", 12.5])

    wb.save(path)
    return path


def test_loads_all_present_sheets_faithfully(tmp_path: Path) -> None:
    path = _build_workbook(tmp_path / "wb.xlsm", pm_header_offset=2)
    loaded = load_workbook(path)

    # Summary: faithful raw grid — 3 rows (incl. its own header row) x 3 cols.
    assert loaded.summary.shape == (3, 3)

    # Profit Margin: header discovered past the 2 junk rows; all 4 data rows kept.
    assert list(loaded.profit_margin.columns) == _PM_HEADER
    assert loaded.profit_margin.shape == (len(_PM_ROWS), len(_PM_HEADER))

    # No rows dropped or altered: both periods present for SKU-A, the
    # zero-activity SKU-B row survives, and the blank row survives.
    pm = loaded.profit_margin
    assert (pm["Marketplace SKU"] == "SKU-A").sum() == 2
    zero_row = pm[(pm["Marketplace SKU"] == "SKU-B") & (pm["Units"] == 0)]
    assert len(zero_row) == 1
    assert zero_row.iloc[0]["Total Gross Sale"] == 0.0
    # The blank-valued SKU-B row is still there (NaN, not filled/dropped).
    blank_row = pm[(pm["Marketplace SKU"] == "SKU-B") & (pm["Units"].isna())]
    assert len(blank_row) == 1

    # Present optional sheet loaded; the three absent ones are None sentinels.
    assert loaded.unmapped_ads is not None
    assert loaded.unmapped_ads.shape == (1, 2)
    assert loaded.canceled_shipping is None
    assert loaded.unmapped_payout is None
    assert loaded.orders_without_payout is None


def test_header_discovery_at_nonzero_offset(tmp_path: Path) -> None:
    # A bigger offset must still resolve to the same correctly-keyed DataFrame.
    path = _build_workbook(tmp_path / "offset.xlsm", pm_header_offset=5)
    loaded = load_workbook(path)
    assert list(loaded.profit_margin.columns) == _PM_HEADER
    assert loaded.profit_margin.shape == (len(_PM_ROWS), len(_PM_HEADER))


def test_missing_required_sheet_raises(tmp_path: Path) -> None:
    path = _build_workbook(tmp_path / "no_pm.xlsm", include_profit_margin=False)
    with pytest.raises(ValueError, match="missing required sheet"):
        load_workbook(path)


def test_missing_optional_sheet_returns_none(tmp_path: Path) -> None:
    # No data-quality sheet at all → no raise; all four optional fields None.
    path = _build_workbook(tmp_path / "no_dq.xlsm", include_unmapped_ads=False)
    loaded = load_workbook(path)
    assert loaded.unmapped_ads is None
    assert loaded.canceled_shipping is None
    assert loaded.unmapped_payout is None
    assert loaded.orders_without_payout is None
    # Core sheets still loaded fine.
    assert not loaded.summary.empty
    assert not loaded.profit_margin.empty


def test_unfindable_profit_margin_header_raises(tmp_path: Path) -> None:
    # Profit Margin sheet present but with no recognizable header row.
    wb = Workbook()
    summary = wb.active
    summary.title = SUMMARY_SHEET
    summary.append(["Line Item", "04/01/2026 - 04/30/2026"])
    pm = wb.create_sheet(PROFIT_MARGIN_SHEET)
    pm.append(["nonsense", "columns", "here"])
    pm.append(["1", "2", "3"])
    path = tmp_path / "bad_header.xlsm"
    wb.save(path)

    with pytest.raises(ValueError, match="header row"):
        load_workbook(path)
